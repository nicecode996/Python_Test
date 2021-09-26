#! usr/bin/env python3
# coding=utf-8

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'test.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 15):
    ws1.append(range(60))

ws2 = wb.create_sheet(title="Pi2")

ws2['F5-F10'] = 3.14

ws3 = wb.create_sheet(title="Data")
ws3['A8'] = ' '
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
ws4 = wb.create_sheet(title='caobi')
for i in range(1, 31):
    ws4.cell(column=1, row=i).value = 'test'

print(ws3['AA10'].value)
wb.save(filename=dest_filename)
