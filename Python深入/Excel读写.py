#! usr/bin/env python3
# coding=utf-8
from openpyxl import load_workbook

wb = load_workbook(filename='empty_book1.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D5'].value)
for i in range(1, 31):
    print(sheet_ranges.cell(column=100, row=i).value)
